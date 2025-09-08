import json
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==== Column order ====
COLUMNS = [
 'NAV_Date','Valuation_Date','FONDO_TARGET_ASSET','OICR_FONDO_TARGET','OICR','ISIN_OICR',
 'Nome_Fondo_Target','ISIN_Fondo_Target','Tipologia_Strumento','Currency_Fondo_Target',
 'Country_Fondo_Target','Hedging_Strategy_Fondo_Target','Strategia_Fondo_Target','Nome_Asset',
 'Nome_Asset_Sintetico','Area_Geo_Asset','Paese_Asset','Indirizzo_Asset','Currency_Asset',
 'Macrosettore_Attivita_Asset','Settore_Attivita_Asset','Tipologia_Investimento',
 'Investment_Date','Exit_Date','Valuation_Methodology','Realized_Unrealized',
 'Commitment_Fondo_Target','Capitale_Investito_Lordo_Loc_Curr','Distribuzioni_Loc_Curr',
 'FMV_Loc_Curr','TVPI','IRR','Cap_Inv_Fondo_Target','FMV_Fondo_Target','Cap_Inv_Ripartito_FOF',
 'NAV_Ripartito_FOF','Capitale_Investito_Fof','FMV_Fof','Possesso_Entry','EV_Entry',
 'EBITDA_Entry','Margin_Entry','Net_Revenue_Entry','Net_Debt_Entry','FCF_Entry',
 'Net_Result_Entry','EVEbitda_Entry','EVRevenue_Entry','Net_DebtEbitda_Entry',
 'Economics_Reporting_Date','Possesso','EV','LTM_EBITDA','LTM_Margin','LTM_Net_Revenues',
 'LTM_Net_Equity','LTM_Net_Debt','LTM_FCF','LTM_Gross_Profit','LTM_Net_Result',
 'Target_Companys_Debt_Equity_Ratio','Discount_Rate','Beta','Cost_of_Equity','Cost_of_Debt',
 'EVEBITDA','EVRevenue','Net_DebtEbitda','Maturity_Date','Price','Coupon','Spread',
 'Watchlist_position','Ltv','Leverage_Entry','Leverage','Duration','Credit_Sensitivity',
 'Risk_Free_Rate_applied_to_Valuation','Credit_Rating','Credit_Rating_Source','Credit_Spread',
 'Country_Premium','Liquidity_Premium','Other_premia',
 'Total_Discount_Rate_applied_to_Valuation','Area','Real_Estate_Segment'
]

# ==== Column groups & colors ====
COLUMN_GROUPS = {
    "beige": ["NAV_Date","Valuation_Date"],
    "gray": ["FONDO_TARGET_ASSET","OICR_FONDO_TARGET","OICR","ISIN_OICR"],
    "light_blue": [
        "Nome_Fondo_Target","ISIN_Fondo_Target","Tipologia_Strumento",
        "Currency_Fondo_Target","Country_Fondo_Target",
        "Hedging_Strategy_Fondo_Target","Strategia_Fondo_Target"
    ],
    "light_yellow": [
        "Nome_Asset","Nome_Asset_Sintetico","Area_Geo_Asset","Paese_Asset",
        "Indirizzo_Asset","Currency_Asset","Macrosettore_Attivita_Asset",
        "Settore_Attivita_Asset"
    ],
    "pink": [
        "Tipologia_Investimento","Investment_Date","Exit_Date","Valuation_Methodology",
        "Realized_Unrealized","Commitment_Fondo_Target","Capitale_Investito_Lordo_Loc_Curr",
        "Distribuzioni_Loc_Curr","FMV_Loc_Curr","TVPI","IRR",
        "Cap_Inv_Fondo_Target","FMV_Fondo_Target","Cap_Inv_Ripartito_FOF",
        "NAV_Ripartito_FOF","Capitale_Investito_Fof","FMV_Fof"
    ],
    "light_green": [
        "Possesso_Entry","EV_Entry","EBITDA_Entry","Margin_Entry","Net_Revenue_Entry",
        "Net_Debt_Entry","FCF_Entry","Net_Result_Entry","EVEbitda_Entry",
        "EVRevenue_Entry","Net_DebtEbitda_Entry","Economics_Reporting_Date","Possesso","EV",
        "LTM_EBITDA","LTM_Margin","LTM_Net_Revenues","LTM_Net_Equity","LTM_Net_Debt",
        "LTM_FCF","LTM_Gross_Profit","LTM_Net_Result"
    ],
    "gray2": ["Target_Companys_Debt_Equity_Ratio","Discount_Rate","Beta","Cost_of_Equity","Cost_of_Debt"],
    "orange": [
        "EVEBITDA","EVRevenue","Net_DebtEbitda","Maturity_Date","Price","Coupon","Spread",
        "Watchlist_position","Ltv","Leverage_Entry","Leverage","Duration","Credit_Sensitivity"
    ],
    "gray3": [
        "Risk_Free_Rate_applied_to_Valuation","Credit_Rating","Credit_Rating_Source",
        "Credit_Spread","Country_Premium","Liquidity_Premium","Other_premia",
        "Total_Discount_Rate_applied_to_Valuation"
    ],
    "green_final": ["Area","Real_Estate_Segment"]
}

FILL_COLORS = {
    "beige": "DDD0C8",
    "gray": "D9D9D9",
    "light_blue": "DEEDED",
    "light_yellow": "F6E8CC",
    "pink": "F8D7DA",
    "light_green": "D4EED4",
    "gray2": "BFBFBF",
    "orange": "F8CBAD",
    "gray3": "C9C9C9",
    "green_final": "C6EFCE"
}

# ==== Functions ====
def flatten_record(rec: Dict[str, Any]) -> Dict[str, Any]:
    flat = {k: rec.get(k, None) for k in COLUMNS}
    snapshots = rec.get("asset_snapshots")
    snap = snapshots[0] if isinstance(snapshots, list) and snapshots else {}
    for k in COLUMNS:
        if k in snap and snap[k] is not None:
            flat[k] = snap[k]
    return flat

def load_json_records(p: Path) -> List[Dict[str, Any]]:
    with p.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data
    elif isinstance(data, dict):
        return [data]
    else:
        return []

def apply_header_colors(excel_path: Path):
    wb = load_workbook(excel_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]

    for idx, col_name in enumerate(header, 1):
        for group, col_list in COLUMN_GROUPS.items():
            if col_name in col_list:
                ws.cell(row=1, column=idx).fill = PatternFill(
                    start_color=FILL_COLORS[group],
                    end_color=FILL_COLORS[group],
                    fill_type="solid"
                )
                break
    wb.save(excel_path)

# ==== Main ====
def export_excel(out_dir: str, output_filename: str = "portfolio_output.xlsx") -> Path | None:
    """
    Export all JSON records in out_dir to an Excel file with colored headers.
    Returns the path to the generated Excel file, or None if no data was found.
    out_dir is a str representing the folder path.
    """
    out_dir_path = Path(out_dir)
    json_files = sorted(out_dir_path.glob("*.json"))
    all_rows: List[Dict[str, Any]] = []
    for jf in json_files:
        try:
            records = load_json_records(jf)
            for rec in records:
                all_rows.append(flatten_record(rec))
        except Exception as e:
            print(f"Skipping {jf.name} due to error: {e}")
    if not all_rows:
        print(f"No valid JSON records found in {out_dir}")
        return None
    df = pd.DataFrame(all_rows, columns=COLUMNS)
    excel_path = out_dir_path / output_filename
    df.to_excel(excel_path, index=False, sheet_name="Data")
    apply_header_colors(excel_path)
    print(f"✅ Wrote {len(df)} rows to {excel_path} with colored headers")
    return excel_path

# if __name__ == "__main__":
#     export_excel()
